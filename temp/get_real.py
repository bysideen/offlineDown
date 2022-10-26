from openpyxl import Workbook, load_workbook




def split_item_str(a):
    index = a.find('(')
    id = a[0:index]
    temp = a[index+1:-1]
    temp = temp.split(',')
    size = int(temp[0])
    length = int(temp[1])
    count = int(temp[2])
    item = [id,size,length,count]
    return item

def real_time_stock():
    wb = load_workbook('stock.xlsx')
    ws = wb['map']
    data = []


    start_code = ws.cell(2,1).value
    if start_code  == 1:
        for row in range(1,20):
            for col in range(1,20):
                value = ws.cell(row,col).value
                if value != None and isinstance(value,str) == True:
                    if value[0].isdigit():
                        data.append(split_item_str(value))
    else:
        return None
                        
    return data



