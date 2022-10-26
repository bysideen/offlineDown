from openpyxl import Workbook, load_workbook
import re 
import tools
import get_real

#init h_item sheet and hl_item sheet
tools.init()
tools.deal_returned_item()

wb = load_workbook('stock.xlsx')
ws = wb['workData']


#printSheet  def
def printSheet(items):
    for each in items:
        print(each)

# Get work data to list(item)
def getDataFromWorkSheet(ws,attrCount):
    items = []
    attrCount = attrCount - 1
    for row in range(2,1000):
        item = []
        for col in range(0,attrCount+2):
            tag = chr(65 + col) + str(row)
            if ws[tag].value == None:
                break
            item.append(ws[tag].value)
            if col == attrCount:
                items.append(item)
    return items
items = getDataFromWorkSheet(ws,3)


#get  itemInfoDict from ItemInfo_sheet
itemInfo_sheet = wb['itemInfo']
itemInfo = []
itemInfo = getDataFromWorkSheet(itemInfo_sheet,5)
itemInfoDict = {}
for info in itemInfo:
    itemInfoDict[info[0]] = info


# Get id_list
def reSetList(l):
    temp = []
    for each in l:
        if not each in temp:
            temp.append(each)
    return temp

id_list = []
for item in items:
    id_list.append(item[0])
id_list = reSetList(id_list)


#原材列表final_items_list料号、规格整理 
items_list = []
for item_id in id_list:
    temp = []
    for item in items:
        if item_id == item[0]:
            temp.append(item)
    items_list.append(temp)
newitems_list = []
for items in items_list:
    temp = []
    for each in items:
        size_temp = [float(s) for s in re.findall(r'-?\d+\.?\d*', each[2])]
        if len(size_temp) > 1:
            size1 = int(size_temp[0])
            size2 = int(size_temp[1])
            if size1 >1:
                size = size1
            else:
                size = size2
        else:
            size = int(size_temp[0])

        temp.append([each[0],size,each[1]])
    newitems_list.append(temp)

final_items_list = []
for items in newitems_list:
    size_list = []
    item_id = items[0][0]
    for item in items:
        size_list.append(item[1])
    size_list = reSetList(size_list)
    temp = []
    for size in size_list:
        temp.append([item_id,size,0])
        
    for each in temp:
        for item in items:
            if item[1] == each[1]:
                each[2] = each[2] + item[2]
    final_items_list.append(temp)

#calculate rolls  of an item.
temp_final = []
for items in final_items_list:
    temp = []
    for item in items:
        tempItem = []
        sourceLength = itemInfoDict[item[0]][3]
        rolls = item[2] // sourceLength 
        length = item[2] - sourceLength * rolls
        tempItem = [item[0],item[1],item[2],rolls,length,sourceLength]
        temp.append(tempItem)
        
    temp_final.append(temp)
final_items_list = temp_final
        

#calculate stock area
half_length_item_sheet = wb['hl_item']
half_length_item_list = tools.extractData(half_length_item_sheet,4)
half_item_sheet = wb['h_item']
half_item_list = tools.extractData(half_item_sheet,4)

for info in itemInfo:
    info[4] = 0

for hfl in half_length_item_list:
    for info in itemInfo:
        if info[0] == hfl[0]:
            area = hfl[1] * hfl[2] / 1000
            info[4] = info[4] + area

for hf in half_item_list:
    for info in itemInfo:
        if info[0] == hf[0]:
            area = hf[1] * hf[2] / 1000
            info[4] = info[4] + area

real_time_items = get_real.real_time_stock()
if real_time_items != None:
    for rti in real_time_items:
        for info in itemInfo:
            info_id = tools.floatToStr(info[0])
            if rti[0] in info_id:
                area = rti[1] * rti[2] * rti[3] / 1000
                info[4] = info[4] + area
def updateInfo(data,ws):
    ws.delete_rows(2,10000)
    for row in range(2,len(data)+2):
        for col in range(1,6):
            ws.cell(row,col,data[row-2][col-1])
updateInfo(itemInfo,itemInfo_sheet)




#append half_item  and half_length:
now = ws['A1'].value
removed_halfs = []
for items in final_items_list:
    count = 1
    for item in items:
        for hfl in half_length_item_list:
            if item[0] == hfl[0] and item[1] == hfl[1]:
                hfl_str = ' ' + str(hfl[2]) + 'm-' + str(hfl[3].year) + '/' + str(hfl[3].month) + '/' + str(hfl[3].day)
                item.append(hfl_str)
                
                if hfl[2] - item[4] >= 0 :
                    new_hfl = hfl[2] - item[4]
                else:
                    new_hfl = item[5] + hfl[2] - item[4]
                item.append('*****'+str(new_hfl) + 'm')
                hfl[2] = new_hfl
                hfl[3] = now

        for hf in half_item_list:
            if item[0] == hf[0] and count ==1:
                hf_str = ' ' + str(hf[1]) + '*' +  str(hf[2]) + 'm-' + str(hf[3].year) + '/' + str(hf[3].month) + '/' + str(hf[3].day)
                item.append(hf_str)
                removed_halfs.append(hf)
        count = count +1

for each in removed_halfs:
    half_item_list.remove(each)
tools.updateStock(half_item_list,half_item_sheet)
tools.updateStock(half_length_item_list,half_length_item_sheet)




#final_items_list  write to workData sheet
ws.delete_rows(1,1000)

header = ['料号','规格(mm)','生产排配(m)','(卷)','需要半成品(m)','标准长度','半成品','日期','原材','日期']
ws.append(header)

for items in final_items_list:
    for item in items:
        ws.append(item)

wb.save('stock.xlsx')
wb.close()
