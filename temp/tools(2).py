from openpyxl import Workbook, load_workbook
import datetime

wb = load_workbook('stock.xlsx')

# Get work data to list(item)
def extractData(ws,attrCount):
    items = []
    attrCount = attrCount -1
    for row in range(2,3000):
        item = []
        for col in range(0,attrCount+1):
            tag = chr(65 + col) + str(row)
            if ws[tag].value == None:
                break
            item.append(ws[tag].value)
            if col == attrCount:
                items.append(item)
    return items
def printSheet(items):
    for each in items:
        print(each)

def floatToStr(item_id):
    if isinstance(item_id,float):
        return str(int(item_id))
    else:
        return str(item_id)

def idCompletion(item_id):
    itemInfo_sheet = wb['itemInfo']
    itemInfo = extractData(itemInfo_sheet,4)
    for each in itemInfo:
        each[0] = floatToStr(each[0])
    for info in itemInfo:
        if item_id in info[0]:
            return info[0]
    return item_id


def updateStock(data,ws):
    ws.delete_rows(2,10000)
    for row in range(2,len(data)+2):
        for col in range(1,5):
            ws.cell(row,col,data[row-2][col-1])
            if col == 4:
                tag = 'D'+str(row)
                cell = ws[tag]
                cell.number_format = 'YYYY/M/D'






def init():
    halfItem_sheet = wb['h_item']
    halfLengthItem_sheet = wb['hl_item']
    halfItem = extractData(halfItem_sheet,4)
    halfLengthItem = extractData(halfLengthItem_sheet,4)

    for item in halfItem:
        item[0] = floatToStr(item[0])
        item[0] = idCompletion(item[0])

    for item in halfLengthItem:
        item[0] = floatToStr(item[0])
        item[0] = idCompletion(item[0])
    updateStock(halfItem,halfItem_sheet)
    updateStock(halfLengthItem,halfLengthItem_sheet)
    wb.save('stock.xlsx')

def split_item_str(a):
    index = a.find('(')
    id = a[0:index]
    temp = a[index+1:-1]
    temp = temp.split(',')
    
    size = int(temp[0]) 
    length = int(temp[1])
    count = int(temp[2])
    item = [id,size,length,count]
    return item

def deal_returned_item():
    init()
    temp_sheet = wb['temp']
    temp = extractData(temp_sheet,4)
    real_stock_sheet = wb['real_stock']
    real_stock = extractData(real_stock_sheet,4)

    halfItem_sheet = wb['h_item']
    halfItem = extractData(halfItem_sheet,4)

    tempNew = []
    halfItemNew = []

    for each in temp:
        flag = False

        for item in halfItem:
            if floatToStr(each[0]) in item[0] and each[1] == item[1]:
                st =  str(int(each[0])) + '(' + str(int(each[1])) + ',' + str(int(each[2])) + ',1'  + ')'
                real_stock.append(st)
                flag = True

        if flag == False:
            tempNew.append(each)

    for item in halfItem:
        flag = False
        for st in real_stock:
            stock = split_item_str(st)
            if stock[0] in  item[0] and stock[1] == item[1]:
                flag = True
                break
        if flag == False:
            halfItemNew.append(item)

    row = 2
    for each in real_stock:
        real_stock_sheet.cell(row,1,each)
        row = row + 1

    updateStock(tempNew,temp_sheet)
    updateStock(halfItemNew,halfItem_sheet)
    wb.save('stock.xlsx')



        


wb.save('stock.xlsx')


