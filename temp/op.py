from openpyxl import Workbook, load_workbook
import re 

wb = load_workbook('stock.xlsx')
ws = wb['workData']

#printSheet  def
def printSheet(items):
    for each in items:
        print(each)

# Get work data to list(item)
def getDataFromWorkSheet(ws,attrCount):
    items = []
    attrCount = attrCount - 1
    for row in range(2,1000):
        item = []
        for col in range(0,5):
            tag = chr(65 + col) + str(row)
            if ws[tag].value == None:
                break
            item.append(ws[tag].value)
            if col == attrCount:
                items.append(item)
    return items
items = getDataFromWorkSheet(ws,3)


#get  itemInfoDict from ItemInfo_sheet
itemInfo_sheet = wb['itemInfo']
itemInfo = []
itemInfo = getDataFromWorkSheet(itemInfo_sheet,4)
itemInfoDict = {}
for info in itemInfo:
    itemInfoDict[info[0]] = info


# Get id_list
def reSetList(l):
    temp = []
    for each in l:
        if not each in temp:
            temp.append(each)
    return temp

id_list = []
for item in items:
    id_list.append(item[0])
id_list = reSetList(id_list)


#原材列表final_items_list料号、规格整理 
items_list = []
for item_id in id_list:
    temp = []
    for item in items:
        if item_id == item[0]:
            temp.append(item)
    items_list.append(temp)
newitems_list = []
for items in items_list:
    temp = []
    for each in items:
        size  = re.findall(r'\d+',each[2])[0]
        size  = int(size)
        temp.append([each[0],size,each[1]])
    newitems_list.append(temp)

final_items_list = []
for items in newitems_list:
    size_list = []
    item_id = items[0][0]
    for item in items:
        size_list.append(item[1])
    size_list = reSetList(size_list)
    temp = []
    for size in size_list:
        temp.append([item_id,size,0])
        
    for each in temp:
        for item in items:
            if item[1] == each[1]:
                each[2] = int(each[2]) + int(item[2])
    final_items_list.append(temp)

#calculate rolls  of an item.
temp_final = []
for items in final_items_list:
    temp = []
    for item in items:
        tempItem = []
        sourceLength = itemInfoDict[item[0]][3]
        rolls = item[2] // sourceLength 
        length = item[2] - sourceLength * rolls
        tempItem = [item[0],item[1],item[2],rolls,length,sourceLength]
        temp.append(tempItem)
    temp_final.append(temp)
final_items_list = temp_final
        


#final_items_list  write to workData sheet
ws.delete_rows(1,1000)

header = ['料号','规格(mm)','生产排配(m)','(卷)','需要半成品(m)','标准长度']
ws.append(header)

for items in final_items_list:
    for item in items:
        ws.append(item)

wb.save('stock.xlsx')
